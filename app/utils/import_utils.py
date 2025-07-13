from openpyxl import load_workbook
from app.models import Person, Transaction
from app.extensions import db
from app.utils.debt_utils import recalculate_debts
from datetime import datetime

def import_excel_transactions(file_stream):
    wb = load_workbook(file_stream)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        timestamp, item, cost, buyer_name, category, recipients_str = row

        try:
            buyer = Person.query.filter_by(name=buyer_name.strip()).first()
            if not buyer:
                continue

            recipient_names = [r.strip() for r in recipients_str.split(",") if r.strip()]
            recipients = Person.query.filter(Person.name.in_(recipient_names)).all()
            if not recipients:
                continue

            txn = Transaction(
                item_name=item,
                cost=float(cost),
                buyer_id=buyer.id,
                budget_category=category
            )

            # Optional: Set timestamp manually if your model supports it
            if hasattr(txn, "timestamp") and timestamp:
                try:
                    txn.timestamp = datetime.strptime(str(timestamp), "%Y-%m-%d %H:%M:%S")
                except:
                    pass

            txn.recipients.extend(recipients)
            db.session.add(txn)
        except Exception as e:
            print("Error on row:", row, e)

    db.session.commit()
    recalculate_debts()
